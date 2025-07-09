import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Font
import re

class BomExcelSortingApp:
    def __init__(self, root):
        self.root = root
        self.root.title('BOM Excel Sorting')
        self.root.configure(bg='#f0f4f8')
        label_font = ('Microsoft JhengHei', 12)
        entry_font = ('Microsoft JhengHei', 11)
        btn_font = ('Microsoft JhengHei', 11, 'bold')
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', font=btn_font, padding=6)
        style.configure('TLabel', font=label_font, background='#f0f4f8')
        style.configure('TEntry', font=entry_font)
        style.configure('TProgressbar', thickness=18, troughcolor='#e0e7ef', background='#4a90e2')

        self.folder_path = tk.StringVar()
        self.save_folder_path = tk.StringVar()

        # 美化UI：設定視窗大小、置中、標題區塊、分隔線、欄位間距
        self.root.geometry('650x480')
        self.root.resizable(False, False)
        # 標題區塊
        title_label = tk.Label(root, text='BOM Excel Sorting 工具', font=('Microsoft JhengHei', 22, 'bold'), bg='#357abd', fg='white', pady=18)
        title_label.grid(row=0, column=0, columnspan=3, sticky='ew')
        # 分隔線
        sep1 = ttk.Separator(root, orient='horizontal')
        sep1.grid(row=1, column=0, columnspan=3, sticky='ew', pady=(0, 16))
        # 調整欄位順序與間距
        row_base = 2
        ttk.Label(root, text='資料夾路徑:').grid(row=row_base, column=0, padx=18, pady=12, sticky='e')
        ttk.Entry(root, textvariable=self.folder_path, width=46, state='readonly').grid(row=row_base, column=1, padx=10, pady=12)
        ttk.Button(root, text='選擇', command=self.select_folder).grid(row=row_base, column=2, padx=10, pady=12)

        ttk.Label(root, text='另存資料夾路徑:').grid(row=row_base+1, column=0, padx=18, pady=12, sticky='e')
        ttk.Entry(root, textvariable=self.save_folder_path, width=46, state='readonly').grid(row=row_base+1, column=1, padx=10, pady=12)
        ttk.Button(root, text='選擇', command=self.select_save_folder).grid(row=row_base+1, column=2, padx=10, pady=12)

        ttk.Label(root, text='主料號:').grid(row=row_base+2, column=0, padx=18, pady=12, sticky='e')
        self.main_part = ttk.Entry(root, width=46)
        self.main_part.grid(row=row_base+2, column=1, padx=10, pady=12, columnspan=2)

        ttk.Label(root, text='替代料:').grid(row=row_base+3, column=0, padx=18, pady=12, sticky='e')
        self.alt_part = ttk.Entry(root, width=46)
        self.alt_part.grid(row=row_base+3, column=1, padx=10, pady=12, columnspan=2)

        ttk.Label(root, text='替代料品名:').grid(row=row_base+4, column=0, padx=18, pady=12, sticky='e')
        self.alt_part_name = ttk.Entry(root, width=46)
        self.alt_part_name.grid(row=row_base+4, column=1, padx=10, pady=12, columnspan=2)

        ttk.Label(root, text='替代料規格:').grid(row=row_base+5, column=0, padx=18, pady=12, sticky='e')
        self.alt_part_spec = ttk.Entry(root, width=46)
        self.alt_part_spec.grid(row=row_base+5, column=1, padx=10, pady=12, columnspan=2)

        # 再加一條分隔線
        sep2 = ttk.Separator(root, orient='horizontal')
        sep2.grid(row=row_base+6, column=0, columnspan=3, sticky='ew', pady=(16, 16))

        # 進度條
        self.progress = ttk.Progressbar(root, orient="horizontal", length=420, mode="determinate", style='TProgressbar')
        self.progress.grid(row=row_base+7, column=0, columnspan=3, pady=20)
        self.progress.grid_remove()

        # 按鈕區塊
        btn_frame = tk.Frame(root, bg='#eaf1fb')
        btn_frame.grid(row=row_base+8, column=0, columnspan=3, pady=16)
        ttk.Button(btn_frame, text='確認', command=self.confirm, width=16).pack(side='left', padx=16)
        ttk.Button(btn_frame, text='取消', command=self.cancel, width=16).pack(side='left', padx=16)
        ttk.Button(btn_frame, text='關閉', command=self.close_app, width=16).pack(side='left', padx=16)

    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)

    def select_save_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.save_folder_path.set(folder_selected)

    def copy_row_format(self, ws, src_row, dst_row):
        # 複製 src_row 的格式到 dst_row
        for col in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            if src_cell.has_style:
                dst_cell._style = src_cell._style  # 複製儲存格樣式
            dst_cell.number_format = src_cell.number_format  # 複製數字格式

    def shift_merged_cells(self, ws, insert_row):
        # 插入行時，調整所有受影響的合併儲存格範圍
        old_ranges = list(ws.merged_cells.ranges)
        ws.merged_cells.ranges = []  # 先清空所有合併範圍

        for merged in old_ranges:
            min_row, max_row = merged.min_row, merged.max_row
            min_col, max_col = merged.min_col, merged.max_col

            if min_row >= insert_row:
                # 整塊合併區在插入點下方，整個往下移動一行
                ws.merge_cells(
                    start_row=min_row + 1,
                    start_column=min_col,
                    end_row=max_row + 1,
                    end_column=max_col
                )
            elif max_row >= insert_row > min_row:
                # 插入點在合併區中間，這塊合併區會被打斷，不再合併
                continue
            else:
                # 不受影響的合併區，保留原樣
                ws.merge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col
                )

    def process_excels(self, folder, main_part, alt_part):
        # 處理所有 Excel 檔案，尋找主料號並插入替代料
        alt_part_name = self.alt_part_name.get()
        alt_part_spec = self.alt_part_spec.get()
        save_folder = self.save_folder_path.get()

        # 收集所有 Excel 檔案
        excel_files = []
        for root_dir, dirs, files in os.walk(folder):
            for file in files:
                if file.lower().endswith('.xlsx') and not file.startswith('~$'):
                    excel_files.append(os.path.join(root_dir, file))

        total = len(excel_files)
        if total == 0:
            return

        self.progress['maximum'] = total
        self.progress['value'] = 0
        self.progress.grid()

        log_path = os.path.join(os.getcwd(), "process_log.txt")
        with open(log_path, 'a', encoding='utf-8') as log_file:
            for idx, file_path in enumerate(excel_files, 1):
                try:
                    wb = load_workbook(file_path)
                    modified = False
                    for ws in wb.worksheets:
                        max_row = ws.max_row
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value == main_part:
                                    insert_row = cell.row + 1  # 要插入的行
                                    found_duplicate = False
                                    check_row = insert_row
                                    # 檢查下方到下一個主料號前有無相同替代料
                                    while check_row <= max_row:
                                        a1_value = ws.cell(row=check_row, column=1).value
                                        alt_value = ws.cell(row=check_row, column=cell.column).value
                                        if a1_value:
                                            break  # 遇到下一個主料號就停止
                                        if alt_value == alt_part:
                                            found_duplicate = True
                                            break
                                        check_row += 1
                                    if not found_duplicate:
                                        ws.insert_rows(insert_row)  # 插入新行
                                        self.shift_merged_cells(ws, insert_row)  # 調整合併儲存格
                                        self.copy_row_format(ws, insert_row - 1, insert_row)  # 複製格式

                                        # 寫入替代料資料
                                        ws.cell(row=insert_row, column=cell.column, value=alt_part)
                                        ws.cell(row=insert_row, column=cell.column + 1, value=alt_part_name)
                                        ws.cell(row=insert_row, column=cell.column + 2, value=alt_part_spec)
                                        ws.cell(row=insert_row, column=cell.column + 11,
                                                value=datetime.datetime.now().strftime('%Y-%m-%d') + ' 客戶確認可替代')

                                        # 設定字體
                                        font_normal = Font(name='MingLiU', size=8)
                                        font_red = Font(name='MingLiU', size=8, color="FFFF0000", bold=True)
                                        ws.cell(row=insert_row, column=cell.column).font = font_normal
                                        ws.cell(row=insert_row, column=cell.column + 1).font = font_normal
                                        ws.cell(row=insert_row, column=cell.column + 2).font = font_normal
                                        ws.cell(row=insert_row, column=cell.column + 11).font = font_red
                                        modified = True

                    if modified:
                        if save_folder:
                            base = os.path.basename(file_path)
                            name, ext = os.path.splitext(base)
                            date_str = datetime.datetime.now().strftime('%Y%m%d')
                            m = re.search(r'(--|-)\d{8}$', name)
                            new_name = re.sub(r'(--|-)\d{8}$', f"{m.group(1)}{date_str}", name) if m else f"{name}-{date_str}"
                            save_path = os.path.join(save_folder, new_name + ext)
                            wb.save(save_path)
                        else:
                            wb.save(file_path)
                    log_file.write(f"{datetime.datetime.now()} - {file_path}\n")
                except Exception as e:
                    log_file.write(f"{datetime.datetime.now()} - {file_path} - Error: {e}\n")

                self.progress['value'] = idx
                self.root.update_idletasks()

        self.progress.grid_remove()
        messagebox.showinfo("完成", "所有檔案處理完成！")

    def confirm(self):
        folder = self.folder_path.get()
        main_part = self.main_part.get()
        alt_part = self.alt_part.get()
        if not folder or not main_part or not alt_part:
            messagebox.showwarning('警告', '請填寫所有欄位')
            return
        self.process_excels(folder, main_part, alt_part)

    def cancel(self):
        self.folder_path.set("")
        self.main_part.delete(0, tk.END)
        self.alt_part.delete(0, tk.END)
        self.alt_part_name.delete(0, tk.END)
        self.alt_part_spec.delete(0, tk.END)
        self.progress.grid_remove()

    def close_app(self):
        self.root.destroy()

if __name__ == '__main__':
    root = tk.Tk()
    app = BomExcelSortingApp(root)
    root.mainloop()
